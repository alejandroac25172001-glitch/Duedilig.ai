import io
from datetime import datetime

import asyncio
from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.schemas import (
    BatchAnalysisResponse,
    DocumentResult,
    ExportRequest,
    GlobalSummary,
)
from app.services.ai_service import analyze_document

router = APIRouter()

# ── colour palette ────────────────────────────────────────────────────────────

_FILL = {
    "ALTO":   PatternFill("solid", fgColor="FFB3B3"),   # rojo claro
    "MEDIO":  PatternFill("solid", fgColor="FFEDB3"),   # amarillo claro
    "BAJO":   PatternFill("solid", fgColor="B3FFB3"),   # verde claro
    "ERROR":  PatternFill("solid", fgColor="D9D9D9"),   # gris
    "HEADER": PatternFill("solid", fgColor="1F4E79"),   # azul oscuro
    "TITLE":  PatternFill("solid", fgColor="2E75B6"),   # azul medio
}

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_WHITE_BOLD = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_DARK_BOLD  = Font(name="Calibri", bold=True, color="1F1F1F", size=11)
_NORMAL     = Font(name="Calibri", color="1F1F1F", size=10)
_WRAP       = Alignment(wrap_text=True, vertical="top")
_CENTER     = Alignment(horizontal="center", vertical="center")


# ── helpers ───────────────────────────────────────────────────────────────────

async def _safe_analyze(file: UploadFile) -> dict:
    """Wraps analyze_document so one failure never aborts the batch."""
    filename = file.filename or "sin_nombre.pdf"
    try:
        pdf_bytes = await file.read()
        if not pdf_bytes:
            raise ValueError("El archivo está vacío.")
        return await analyze_document(pdf_bytes, filename)
    except Exception as exc:
        return {
            "filename": filename,
            "tiene_autorizacion": False,
            "menciona_centrales_riesgo": False,
            "nivel_riesgo": "ERROR",
            "fragmento_relevante": None,
            "resumen": f"Error al procesar el documento: {exc}",
            "confianza": 0,
        }


def _compute_summary(resultados: list[dict]) -> GlobalSummary:
    total    = len(resultados)
    alto     = sum(1 for r in resultados if r["nivel_riesgo"] == "ALTO")
    medio    = sum(1 for r in resultados if r["nivel_riesgo"] == "MEDIO")
    bajo     = sum(1 for r in resultados if r["nivel_riesgo"] == "BAJO")
    errores  = sum(1 for r in resultados if r["nivel_riesgo"] == "ERROR")
    analizados = total - errores
    pct = round(bajo / analizados * 100, 1) if analizados > 0 else 0.0
    return GlobalSummary(
        total=total,
        alto=alto,
        medio=medio,
        bajo=bajo,
        errores=errores,
        porcentaje_cumplimiento=pct,
    )


# ── Excel builder ─────────────────────────────────────────────────────────────

def _style_cell(cell, fill=None, font=None, alignment=None, border=True):
    if fill:      cell.fill      = fill
    if font:      cell.font      = font
    if alignment: cell.alignment = alignment
    if border:    cell.border    = _THIN_BORDER


def _build_excel(resultados: list[dict], resumen: GlobalSummary) -> bytes:
    wb = Workbook()

    # ── Hoja 1: Resumen ───────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.column_dimensions["A"].width = 36
    ws1.column_dimensions["B"].width = 20

    # Título
    ws1.merge_cells("A1:B1")
    title_cell = ws1["A1"]
    title_cell.value = "DUEDILIG-AI — Resumen de Análisis"
    title_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    title_cell.fill = _FILL["TITLE"]
    title_cell.alignment = _CENTER
    ws1.row_dimensions[1].height = 30

    # Sub-título con fecha
    ws1.merge_cells("A2:B2")
    date_cell = ws1["A2"]
    date_cell.value = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    date_cell.font  = Font(name="Calibri", italic=True, color="595959", size=10)
    date_cell.alignment = _CENTER

    # Encabezado de tabla
    for col, label in [(1, "Métrica"), (2, "Valor")]:
        c = ws1.cell(row=4, column=col, value=label)
        _style_cell(c, fill=_FILL["HEADER"], font=_WHITE_BOLD, alignment=_CENTER)
    ws1.row_dimensions[4].height = 20

    summary_rows = [
        ("Total de documentos analizados",  resumen.total,                    None),
        ("Riesgo ALTO (sin autorización)",   resumen.alto,                     "ALTO"),
        ("Riesgo MEDIO (autorización genérica)", resumen.medio,                "MEDIO"),
        ("Riesgo BAJO (cumple normativa)",   resumen.bajo,                     "BAJO"),
        ("Errores de procesamiento",         resumen.errores,                  "ERROR"),
        ("% de cumplimiento normativo",      f"{resumen.porcentaje_cumplimiento}%", "BAJO" if resumen.porcentaje_cumplimiento >= 50 else "ALTO"),
    ]

    for row_idx, (label, value, risk) in enumerate(summary_rows, start=5):
        fill = _FILL.get(risk) if risk else None

        label_cell = ws1.cell(row=row_idx, column=1, value=label)
        _style_cell(label_cell, fill=fill, font=_DARK_BOLD, alignment=Alignment(vertical="center", indent=1))

        value_cell = ws1.cell(row=row_idx, column=2, value=value)
        _style_cell(value_cell, fill=fill, font=_DARK_BOLD, alignment=_CENTER)
        ws1.row_dimensions[row_idx].height = 20

    # ── Hoja 2: Detalle ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Detalle")

    col_widths = [5, 40, 22, 22, 14, 12, 60, 60]
    for i, w in enumerate(col_widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    headers = [
        "#",
        "Archivo",
        "Tiene Autorización",
        "Menciona Centrales",
        "Nivel de Riesgo",
        "Confianza (%)",
        "Fragmento Relevante",
        "Resumen",
    ]
    for col, label in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=col, value=label)
        _style_cell(c, fill=_FILL["HEADER"], font=_WHITE_BOLD, alignment=_CENTER)
    ws2.row_dimensions[1].height = 24

    for row_idx, doc in enumerate(resultados, start=2):
        risk   = doc.get("nivel_riesgo", "ERROR")
        fill   = _FILL.get(risk)

        values = [
            row_idx - 1,
            doc.get("filename", ""),
            "Sí" if doc.get("tiene_autorizacion") else "No",
            "Sí" if doc.get("menciona_centrales_riesgo") else "No",
            risk,
            doc.get("confianza", 0),
            doc.get("fragmento_relevante") or "—",
            doc.get("resumen", ""),
        ]
        for col, value in enumerate(values, start=1):
            c = ws2.cell(row=row_idx, column=col, value=value)
            _style_cell(c, fill=fill, font=_NORMAL, alignment=_WRAP)

        ws2.row_dimensions[row_idx].height = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── endpoints ─────────────────────────────────────────────────────────────────

@router.post("/upload", response_model=BatchAnalysisResponse)
async def upload_contracts(files: list[UploadFile] = File(...)):
    """
    Recibe uno o varios PDFs, los analiza en paralelo y devuelve los resultados
    junto con un resumen global de cumplimiento normativo.
    """
    if not files:
        raise HTTPException(status_code=422, detail="Debe enviar al menos un archivo PDF.")

    raw_results: list[dict] = await asyncio.gather(
        *[_safe_analyze(f) for f in files]
    )

    resultados  = [DocumentResult(**r) for r in raw_results]
    resumen     = _compute_summary(raw_results)

    return BatchAnalysisResponse(resultados=resultados, resumen_global=resumen)


@router.post("/export/excel")
async def export_excel(body: ExportRequest):
    """
    Recibe la lista de resultados y devuelve un archivo Excel descargable
    con dos hojas: Resumen (métricas con colores) y Detalle (una fila por documento).
    """
    if not body.resultados:
        raise HTTPException(status_code=422, detail="La lista de resultados está vacía.")

    raw = [r.model_dump() for r in body.resultados]
    resumen = _compute_summary(raw)
    excel_bytes = _build_excel(raw, resumen)

    filename = f"duedilig_reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
